import os
import openpyxl
import csv

def is_row_empty(row):
    """Check if a row is completely empty (None or empty string after stripping)"""
    return all(cell is None or str(cell).strip() == "" for cell in row)

def clean_row(row):
    """Clean a row by removing trailing empty cells and cleaning values"""
    # Convert None to empty string and strip whitespace
    cleaned = [str(cell).strip() if cell is not None else "" for cell in row]
    # Remove trailing empty cells
    while len(cleaned) > 0 and cleaned[-1] == "":
        cleaned.pop()
    return cleaned

def excel_to_csv(input_file, output_folder):
    """
    Convert Excel to CSV while:
    1. Removing completely empty rows
    2. Removing trailing empty cells (reduces commas)
    3. Stripping whitespace from cells
    """
    os.makedirs(output_folder, exist_ok=True)
    workbook = openpyxl.load_workbook(input_file)
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        safe_name = "".join(c if c.isalnum() else "_" for c in sheet_name)
        csv_path = os.path.join(output_folder, f"{safe_name}.csv")
        
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Process all rows
            rows_to_write = []
            max_cols = 0
            
            for row in sheet.iter_rows(values_only=True):
                if not is_row_empty(row):  # Skip empty rows
                    cleaned = clean_row(row)
                    rows_to_write.append(cleaned)
                    if len(cleaned) > max_cols:
                        max_cols = len(cleaned)
            
            # Write all non-empty rows with consistent columns
            for row in rows_to_write:
                writer.writerow(row + [""] * (max_cols - len(row)))
        
        original_size = sheet.max_row
        cleaned_size = len(rows_to_write)
        print(f"Saved {sheet_name}: {original_size} → {cleaned_size} rows (-{100*(original_size-cleaned_size)/original_size:.1f}%)")

if __name__ == "__main__":
    excel_to_csv("data/test-large.xlsx", "sheets")